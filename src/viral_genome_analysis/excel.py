from __future__ import annotations

from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .utils import safe_sheet_name

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
INPUT_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
FLAG_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
SIGNIFICANT_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")


def _apply_table(ws, max_row: int, max_col: int, table_name: str) -> None:
    if max_row < 2 or max_col < 1:
        return
    end_col = get_column_letter(max_col)
    table = Table(displayName=table_name, ref=f"A1:{end_col}{max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def _format_numbers(ws) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}
    for header, column in headers.items():
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2, max_row=ws.max_row):
            for entry in cell:
                if entry.value is None:
                    continue
                if isinstance(entry.value, float):
                    if "p_value" in str(header):
                        entry.number_format = "0.0000"
                    elif "fraction" in str(header) or "accuracy" in str(header) or "r_squared" in str(header) or "silhouette" in str(header):
                        entry.number_format = "0.000"
                    else:
                        entry.number_format = "0.000"
                elif isinstance(entry.value, int):
                    entry.number_format = "0"


def _highlight_special_columns(ws) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}
    p_cols = [col for name, col in headers.items() if "p_value" in str(name)]
    for row in range(2, ws.max_row + 1):
        for col in p_cols:
            value = ws.cell(row=row, column=col).value
            try:
                if value is not None and float(value) <= 0.05:
                    ws.cell(row=row, column=col).fill = SIGNIFICANT_FILL
            except Exception:
                continue


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    for col_cells in ws.columns:
        max_length = 0
        column_index = col_cells[0].column
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 42)
    _format_numbers(ws)
    _highlight_special_columns(ws)


def write_workbook(sheets: Dict[str, pd.DataFrame], output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    table_counter = 1
    for sheet_name, df in sheets.items():
        ws = workbook.create_sheet(safe_sheet_name(sheet_name))
        export_df = df.copy()
        if export_df.empty:
            export_df = pd.DataFrame({"note": ["No rows available"]})
        ws.append(list(export_df.columns))
        for row in export_df.itertuples(index=False, name=None):
            ws.append(list(row))
        _style_sheet(ws)
        _apply_table(ws, ws.max_row, ws.max_column, f"Table{table_counter}")
        table_counter += 1

    workbook.save(output_path)
    return output_path


def write_csv_bundle(sheets: Dict[str, pd.DataFrame], output_dir: str | Path) -> None:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    for name, df in sheets.items():
        export_path = output_dir / f"{name}.csv"
        df.to_csv(export_path, index=False)
